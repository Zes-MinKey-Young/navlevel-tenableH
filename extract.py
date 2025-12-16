from openpyxl import load_workbook

from typing import TypedDict

import json

wb = load_workbook("tenable/ten.xlsx", data_only=True)


class Level(TypedDict):
    name: str
    score: float
    beau: str
    crea: str
    expr: str
    play: str
    logic: str


sheet = wb.active
if not sheet:
    print("No active sheet found")
    exit(1)

levels: list[Level] = []

pos = 0

for row in sheet.iter_rows(values_only=True):
    count = len(row)
    if pos < 2:
        pos += 1
        continue
    if row[0] is None:
        break
    for i in range(count):
        print(i, row[i], end=" ")
    levels.append({
        "name": row[0],
        "score": row[1],
        "beau": f"{row[5]} ({row[19]})",
        "crea": f"{row[8]} ({row[20]})",
        "expr": f"{row[11]} ({row[21]})",
        "play": f"{row[14]} ({row[23]})",
        "logic": f"{row[17]} ({row[23]})"
    })

json.dump(levels, open("tenable/ten.json", "w", encoding="utf-8"), ensure_ascii=False, indent=4)